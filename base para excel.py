from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.chart import PieChart, BarChart, LineChart, Reference

wb = Workbook()

# ==========================
# ABA 1 - PARCELAMENTOS
# ==========================

ws1 = wb.active
ws1.title = "Parcelamentos Ativos"

headers = [
    "Loja",
    "Valor Parcela (R$)",
    "Parcelas Restantes",
    "Total Futuro (R$)",
    "Categoria"
]

for col, h in enumerate(headers, start=1):
    cell = ws1.cell(row=1, column=col)
    cell.value = h
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1F4E78")

dados = [
    ["MSC Cruzeiros",890.87,11,9799.57,"Lazer/Viagem"],
    ["Adidas",93.04,9,837.36,"Vestuário"],
    ["Leroy Merlin",259.12,2,518.24,"Casa"],
    ["Distmachado",333.33,3,999.99,"Casa"],
    ["Lojas Renner",146.60,2,293.20,"Vestuário"],
    ["Santa Lolla",66.64,2,133.28,"Vestuário"],
    ["Importsg10",347.22,12,4166.64,"Eletrônicos"],
    ["Green Tecnologia",207.75,9,1869.75,"Tecnologia"],
    ["Refinanciamento Fatura",2082.79,6,12496.74,"Financeiro"]
]

for row in dados:
    ws1.append(row)

# ==========================
# ABA 2 - COMPRAS À VISTA
# ==========================

ws2 = wb.create_sheet("Compras à Vista")

headers2 = [
    "Data",
    "Descrição",
    "Valor (R$)",
    "Categoria"
]

for col, h in enumerate(headers2, start=1):
    cell = ws2.cell(row=1, column=col)
    cell.value = h
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="38761D")

dados2 = [
    ["16/04","Netflix",44.90,"Lazer"],
    ["18/04","Burger King",46.90,"Alimentação"],
    ["19/04","Mc Donalds",65.90,"Alimentação"],
    ["25/04","Allianz Seguro",72.48,"Serviços"],
    ["28/04","Asaás FisioSaúde",420.00,"Saúde"],
    ["04/05","Bellavia",214.59,"Alimentação"],
    ["09/05","Santini Esmalteria",185.00,"Beleza"],
    ["10/05","Panificadora Sabor",46.25,"Alimentação"]
]

for row in dados2:
    ws2.append(row)

# ==========================
# ABA 3 - RESUMO
# ==========================

ws3 = wb.create_sheet("Resumo Mensal")

headers3 = [
    "Mês",
    "Total Parcelamentos",
    "Total Compras à Vista",
    "Total Fatura Prevista"
]

for col, h in enumerate(headers3, start=1):
    cell = ws3.cell(row=1, column=col)
    cell.value = h
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="C00000")

dados3 = [
    ["Maio/2026",3952,1200,5152],
    ["Junho/2026",3952,800,4752],
    ["Julho/2026",3952,950,4902]
]

for row in dados3:
    ws3.append(row)

# ==========================
# DASHBOARD
# ==========================

dash = wb.create_sheet("Dashboard")

dash["A1"] = "CONTROLE FINANCEIRO"
dash["A1"].font = Font(size=16, bold=True)

dash["A3"] = "Limite Cartão"
dash["B3"] = 20000

dash["A4"] = "Utilizado"
dash["B4"] = 16114.03

dash["A5"] = "Disponível"
dash["B5"] = "=B3-B4"

# ==========================
# GRÁFICO DE LINHA
# ==========================

line = LineChart()
line.title = "Evolução das Faturas"

data = Reference(ws3, min_col=4, min_row=1, max_row=4)
cats = Reference(ws3, min_col=1, min_row=2, max_row=4)

line.add_data(data, titles_from_data=True)
line.set_categories(cats)

dash.add_chart(line, "E8")

# ==========================
# GRÁFICO DE BARRAS
# ==========================

bar = BarChart()
bar.title = "À Vista x Parcelado"

data = Reference(ws3, min_col=2, max_col=3, min_row=1, max_row=4)
cats = Reference(ws3, min_col=1, min_row=2, max_row=4)

bar.add_data(data, titles_from_data=True)
bar.set_categories(cats)

dash.add_chart(bar, "E25")

# ==========================
# TABELA CATEGORIAS
# ==========================

dash["M2"] = "Categoria"
dash["N2"] = "Valor"

categorias = [
    ["Lazer/Viagem",9799.57],
    ["Vestuário",1263.84],
    ["Casa",1518.23],
    ["Eletrônicos",4166.64],
    ["Tecnologia",1869.75],
    ["Financeiro",12496.74]
]

for linha in categorias:
    dash.append(linha)

# ==========================
# GRÁFICO PIZZA
# ==========================

pie = PieChart()
pie.title = "Gastos por Categoria"

labels = Reference(dash, min_col=13, min_row=3, max_row=8)
data = Reference(dash, min_col=14, min_row=2, max_row=8)

pie.add_data(data, titles_from_data=True)
pie.set_categories(labels)

dash.add_chart(pie, "M10")

# ==========================
# FORMATAÇÃO MONETÁRIA
# ==========================

for ws in wb.worksheets:
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = 'R$ #,##0.00'

# ==========================
# SALVAR
# ==========================

wb.save("Controle_Financeiro_Debora.xlsx")

print("Arquivo gerado com sucesso!")